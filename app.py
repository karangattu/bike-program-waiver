import base64
import io
import os
import textwrap
from datetime import datetime
from urllib.parse import urlparse

import htmltools as tags
import requests
from openpyxl import Workbook
from PIL import Image, ImageDraw, ImageFont
from shiny import App, reactive, render, ui

try:
    from dotenv import load_dotenv

    load_dotenv()
except Exception:
    pass

def ensure_fonts_available():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    fonts_dir = os.path.join(script_dir, "fonts")
    
    os.makedirs(fonts_dir, exist_ok=True)
    
    font_urls = {
        "NotoSansCJKsc-Regular.otf": "https://github.com/googlefonts/noto-cjk/raw/main/Sans/OTF/SimplifiedChinese/NotoSansCJKsc-Regular.otf",
        "NotoSansCJKsc-Bold.otf": "https://github.com/googlefonts/noto-cjk/raw/main/Sans/OTF/SimplifiedChinese/NotoSansCJKsc-Bold.otf",
        "OpenSans-Regular.ttf": "https://github.com/google/fonts/raw/main/apache/opensans/OpenSans%5Bwdth%2Cwght%5D.ttf"
    }
    
    for font_name, font_url in font_urls.items():
        font_path = os.path.join(fonts_dir, font_name)
        if not os.path.exists(font_path):
            try:
                print(f"[fonts] Downloading {font_name}...")
                response = requests.get(font_url, timeout=30)
                response.raise_for_status()
                with open(font_path, 'wb') as f:
                    f.write(response.content)
                print(f"[fonts] Downloaded {font_name} successfully")
            except Exception as e:
                print(f"[fonts] Failed to download {font_name}: {e}")

try:
    ensure_fonts_available()
except Exception as e:
    print(f"[fonts] Font initialization error: {e}")

waiver_content = {
    "en": {
        "title": "BICYCLE PROGRAM WAIVER AND RELEASE FROM LIABILITY",
        "logo_text1": "HOPE'S CORNER",
        "logo_subtext": "SHARING MEALS, BUILDING COMMUNITY",
        "intro": "I expressly assume and accept any and all risk of injury or death to myself or others arising from my use of the Hope's Corner, Inc. Bicycle Program including repair services, transportation equipment such as bicycles, scooters, skateboards, etc. and all equipment and supplies. This waiver and release from liability includes any and all repair services, equipment, including but not limited to, the bicycle, helmet, lock, light, rack, basket, tubes, tires, chains, brakes, and my participation in the Program.",
        "points": [
            "I am a voluntary participant and utilize the program services and equipment at my own risk.",
            "I am responsible for maintaining the bicycle in good condition. I will inspect the bicycle prior to use to ensure all parts are in proper working condition.",
            "I understand that there are risks inherent in riding a bicycle even when the bicycle and equipment are in good working order and used properly. Injuries are a common, ordinary and foreseeable consequence of bicycle riding. I understand that the risks I may encounter include, but are not limited to the following:",
        ],
        "sub_points": [
            "The equipment may break or malfunction, causing loss of or damage to property or injury to my person or to another person.",
            "Riding a bicycle requires physical exertion and may result in discomfort, pain or injury.",
            "I might encounter hazards while riding which could cause me to fall or be propelled.",
            "Bicycle travel is dangerous. I may be injured by many factors over which I have no control.",
        ],
        "volunteer_clause": "I understand that I assume the risk of any damage or failure of equipment arising from repair work or services donated to me by the Bicycle Program volunteers, employees or vendors. I understand the program volunteers and employees are not trained or licensed professionals and cannot be held responsible for damage to my equipment. I am responsible for inspecting the work and not utilizing the bicycle or equipment if it is not in good working order.",
        "indemnify_clause": "I agree to indemnify, defend, save and hold harmless Hope's Corner, Inc. including all employees, volunteers, directors, officers, vendors and funders of the program from any claims, losses, damages or liability accruing or resulting to any person or entity from my participation in the Bicycle Program or use of the bicycle, bicycle equipment and bicycle repair services.",
        "release_clause": "I, on behalf of myself, my heirs, successors and assigns, hereby waive, release and forever discharge Hope's Corner, Inc. including all employees, volunteers, directors, officers, vendors and funders of the program from any and all claims, losses, damages or liability accruing or resulting to any person or entity from my participation in the Bicycle Program or use of the bicycle, bicycle equipment and repair services.",
        "final_agreement": "I AM AWARE THAT THIS IS A RELEASE OF LIABILITY. I AM SIGNING IT FREELY AND OF MY OWN ACCORD AND I RECOGNIZE AND AGREE THAT IT IS BINDING UPON MYSELF, MY HEIRS AND ASSIGNS, AND IN THE EVENT THAT I AM SIGNING IT ON BEHALF OF ANY MINORS, I HAVE FULL LEGAL AUTHORITY TO DO SO, AND REALIZE THE BINDING EFFECT OF THIS CONTRACT ON THEM, AS WELL AS ON MYSELF. I AGREE TO ALLOW HOPE'S CORNER, INC. TO USE PHOTOGRAPHS, VIDEOS, OR SOUND RECORDINGS OF ME FOR PROMOTIONAL AND PUBLICITY PURPOSES.",
        "agreement_check": "I have carefully read this agreement and understand that this is a release of liability. I agree to the terms and conditions.",
        "print_name_label": "Print Full Name",
        "signature_label": "Signature (Draw below)",
        "date_label": "Date",
        "submit_button": "Submit Waiver",
        "switch_button": "Español | 中文",
        "success_message": "Thank you! Your waiver has been submitted successfully.",
        "signature_placeholder": "Type your signature here or upload an image",
    },
    "es": {
        "title": "PROGRAMA DE BICICLETAS ACUERDO DE RENUNCIA Y EXENCIÓN DE RESPONSABILIDAD",
        "logo_text1": "HOPE'S CORNER",
        "logo_subtext": "COMPARTIR COMIDAS, CONSTRUYENDO COMUNIDAD",
        "intro": "Asumo y acepto expresamente todos y cada uno de los riesgos de lesiones o muerte para mí u otros que surjan de mi uso del Programa de Bicicletas de Hope's Corner, Inc., incluidos los servicios de reparación, equipos de transporte como bicicletas, scooters, patinetas, etc., y todo el equipo y los suministros. Esta renuncia y exención de responsabilidad incluye todos y cada uno de los servicios de reparación, equipos, incluidos, entre otros, la bicicleta, el casco, el candado, la luz, el portabultos, la canasta, las cámaras, los neumáticos, las cadenas, los frenos y mi participación en el Programa.",
        "points": [
            "Soy un participante voluntario y utilizo los servicios y el equipo del programa bajo mi propio riesgo.",
            "Soy responsable de mantener la bicicleta en buenas condiciones. Inspeccionaré la bicicleta antes de usarla para asegurarme de que todas las piezas estén en condiciones de funcionamiento adecuadas.",
            "Entiendo que existen riesgos inherentes al andar en bicicleta incluso cuando la bicicleta y el equipo están en buen estado de funcionamiento y se utilizan correctamente. Las lesiones son una consecuencia común, ordinaria y previsible de andar en bicicleta. Entiendo que los riesgos que puedo encontrar incluyen, entre otros, los siguientes:",
        ],
        "sub_points": [
            "El equipo puede romperse o funcionar mal, causando pérdidas o daños a la propiedad o lesiones a mi persona o a otra persona.",
            "Andar en bicicleta requiere esfuerzo físico y puede provocar molestias, dolor o lesiones.",
            "Podría encontrar peligros mientras conduzco que podrían hacer que me caiga o sea impulsado.",
            "Viajar en bicicleta es peligroso. Puedo resultar herido por muchos factores sobre los que no tengo control.",
        ],
        "volunteer_clause": "Entiendo que asumo el riesgo de cualquier daño o falla del equipo que surja del trabajo de reparación o los servicios que me donen los voluntarios, empleados o proveedores del Programa de Bicicletas. Entiendo que los voluntarios y empleados del programa no son profesionales capacitados o con licencia y no se les puede hacer responsables de los daños a mi equipo. Soy responsible de inspeccionar el trabajo y no utilizar la bicicleta o el equipo si no está en buen estado de funcionamiento.",
        "indemnify_clause": "Acepto indemnizar, defender, salvar y mantener indemne a Hope's Corner, Inc., incluidos todos los empleados, voluntarios, directores, funcionarios, proveedores y financiadores del programa de cualquier reclamo, pérdida, daño o responsabilidad que se acumule o resulte para cualquier persona o entidad de mi participación en el Programa de Bicicletas o el uso de la bicicleta, el equipo de la bicicleta y los servicios de reparación de bicicletas.",
        "release_clause": "Yo, en mi nombre, mis herederos, sucesores y cesionarios, por la presente renuncio, libero y descargo para siempre a Hope's Corner, Inc., incluidos todos los empleados, voluntarios, directores, funcionarios, proveedores y financiadores del programa de cualquier y todos los reclamos, pérdidas, daños o responsabilidad que se acumulen o resulten para cualquier persona o entidad de mi participación en el Programa de Bicicletas o el uso de la bicicleta, el equipo de la bicicleta y los servicios de reparación.",
        "final_agreement": "SOY CONSCIENTE DE QUE ESTA ES UNA EXENCIÓN DE RESPONSABILIDAD. LO FIRMO LIBREMENTE Y POR MI PROPIA VOLUNTAD Y RECONOZCO Y ACEPTO QUE ES VINCULANTE PARA MÍ, MIS HEREDEROS Y CESIONARIOS, Y EN CASO DE QUE LO FIRME EN NOMBRE DE MENORES, TENGO PLENA AUTORIDAD LEGAL PARA HACERLO, Y ME DOY CUENTA DEL EFECTO VINCULANTE DE ESTE CONTRATO SOBRE ELLOS, ASÍ COMO SOBRE MÍ. ACEPTO PERMITIR QUE HOPE'S CORNER, INC. UTILICE FOTOGRAFÍAS, VIDEOS O GRABACIONES DE SONIDO DE MÍ PARA FINES PROMOCIONALES Y PUBLICITARIOS.",
        "agreement_check": "He leído atentamente este acuerdo y entiendo que se trata de una exención de responsabilidad. Acepto los términos y condiciones.",
        "print_name_label": "Escriba el Nombre Completo",
        "signature_label": "Firma (Dibuje abajo)",
        "date_label": "Fecha",
        "submit_button": "Enviar Renuncia",
        "switch_button": "English | 中文",
        "success_message": "¡Gracias! Su renuncia ha sido enviada con éxito.",
        "signature_placeholder": "Escriba su firma aquí o suba una imagen",
    },
    "zh": {
        "title": "自行车项目弃权和免责书",
        "logo_text1": "希望之角",
        "logo_subtext": "共享美食，共建社区",
        "intro": "我明确承担并接受因使用希望之角公司自行车项目（包括维修服务、自行车、滑板车、滑板等交通设备以及所有设备和用品）而对我自己或他人造成的任何和所有伤害或死亡风险。本弃权和免责书包括任何和所有维修服务、设备，包括但不限于自行车、头盔、锁、灯、货架、篮子、内胎、轮胎、链条、制动器，以及我对该项目的参与。",
        "points": [
            "我是自愿参与者，自行承担使用项目服务和设备的风险。",
            "我有责任保持自行车处于良好状态。我将在使用前检查自行车，确保所有部件都处于正常工作状态。",
            "我理解骑自行车存在固有风险，即使自行车和设备处于良好工作状态并正确使用。伤害是骑自行车的常见、普通和可预见的后果。我理解可能遇到的风险包括但不限于以下内容：",
        ],
        "sub_points": [
            "设备可能损坏或故障，造成财产损失或损害，或对我本人或他人造成伤害。",
            "骑自行车需要体力消耗，可能导致不适、疼痛或受伤。",
            "我可能在骑行时遇到危险，这可能导致我跌倒或被推动。",
            "自行车出行是危险的。我可能因许多我无法控制的因素而受伤。",
        ],
        "volunteer_clause": "我理解我承担因自行车项目志愿者、员工或供应商为我提供的维修工作或服务而导致的任何设备损坏或故障的风险。我理解项目志愿者和员工不是经过培训或持有执照的专业人员，不能对我的设备损坏负责。我有责任检查工作，如果自行车或设备未处于良好工作状态，则不使用它们。",
        "indemnify_clause": "我同意赔偿、辩护、拯救并使希望之角公司（包括项目的所有员工、志愿者、董事、管理人员、供应商和资助者）免受因我参与自行车项目或使用自行车、自行车设备和自行车维修服务而对任何个人或实体产生或导致的任何索赔、损失、损害或责任的损害。",
        "release_clause": "我代表我本人、我的继承人、继任者和受让人，在此放弃、免除并永远解除希望之角公司（包括项目的所有员工、志愿者、董事、管理人员、供应商和资助者）因我参与自行车项目或使用自行车、自行车设备和自行车维修服务而对任何个人或实体产生或导致的任何和所有索赔、损失、损害或责任。",
        "final_agreement": "我知道这是一份免责书。我是自由和自愿签署的，我认识并同意它对我本人、我的继承人和受让人具有约束力，如果我代表任何未成年人签署，我有充分的法律权力这样做，并意识到本合同对他们以及对我的约束效力。我同意允许希望之角公司将我的照片、视频或录音用于宣传和宣传目的。",
        "agreement_check": "我已仔细阅读本协议，并理解这是一份免责书。我同意条款和条件。",
        "print_name_label": "打印全名",
        "signature_label": "签名（请在下方签名）",
        "date_label": "日期",
        "submit_button": "提交弃权书",
        "switch_button": "English | Español",
        "success_message": "谢谢！您的弃权书已成功提交。",
        "signature_placeholder": "在此输入您的签名或上传图片",
    },
}


def create_header(content, language):
    return tags.div(
        tags.div(
            tags.div(
                ui.input_action_button(
                    "language_switch",
                    content["switch_button"],
                    class_="btn btn-outline-light btn-sm",
                ),
                style="position: absolute; top: 1.5rem; right: 1.5rem; z-index: 10;",
            ),
            tags.div(
                tags.div(
                    tags.img(
                        src="https://images.squarespace-cdn.com/content/v1/5622cd82e4b0501d40689558/cdab4aef-0027-40b7-9737-e2f893586a6a/Hopes_Corner_Logo_Green.png?format=500w",
                        alt="Hope's Corner Logo",
                        style="height: 90px; object-fit: contain; margin-bottom: 1rem;",
                    ),
                    class_="logo-container text-center",
                ),
                tags.h1(content["logo_text1"], class_="h2 fw-bold text-white mb-2"),
                tags.p(content["logo_subtext"], class_="text-white opacity-75 mb-0 fw-medium"),
                class_="text-center",
            ),
            class_="header-section text-dark p-5 position-relative",
        )
    )


def create_waiver_content(content):
    return ui.tags.div(
        ui.tags.h3(content["title"], class_="waiver-title"),
        ui.tags.div(
            ui.tags.p(content["intro"], class_="mb-4"),
            ui.tags.ol(
                ui.tags.li(content["points"][0], class_="mb-3"),
                ui.tags.li(content["points"][1], class_="mb-3"),
                ui.tags.li(
                    content["points"][2],
                    ui.tags.ol(
                        *[
                            ui.tags.li(point, class_="mb-2")
                            for point in content["sub_points"]
                        ],
                        type="a",
                        class_="mt-3 ps-3",
                    ),
                    class_="mb-3",
                ),
                class_="ps-3",
            ),
            ui.tags.hr(class_="divider"),
            tags.p(content["volunteer_clause"], class_="mb-3"),
            tags.p(content["indemnify_clause"], class_="mb-3"),
            tags.p(content["release_clause"], class_="mb-4"),
            tags.div(
                tags.p(
                    content["final_agreement"], class_="fw-bold text-danger mb-0"
                ),
                class_="alert alert-warning border-start border-warning border-4 bg-warning bg-opacity-10 rounded-3",
            ),
            class_="waiver-content",
        ),
    )


def create_form_section(content):
    return tags.div(
        ui.tags.hr(class_="divider"),
        ui.tags.div(
            ui.tags.div(
                ui.input_checkbox("agreement", content["agreement_check"], value=False),
                class_="mb-4",
            ),
            ui.tags.div(
                ui.tags.div(
                    ui.tags.div(
                        ui.input_text(
                            "participant_name",
                            content["print_name_label"],
                            placeholder="Enter your full name",
                        ),
                        ui.tags.script(
                            """
                            document.addEventListener('DOMContentLoaded', function() {
                                function setupNameCapitalization() {
                                    const nameInput = document.querySelector('input[id*="participant_name"]');
                                    if (nameInput) {
                                        nameInput.addEventListener('blur', function() {
                                            const words = this.value.split(' ');
                                            const capitalizedWords = words.map(word => {
                                                if (word.length > 0) {
                                                    return word.charAt(0).toUpperCase() + word.slice(1).toLowerCase();
                                                }
                                                return word;
                                            });
                                            this.value = capitalizedWords.join(' ');
                                            this.dispatchEvent(new Event('input', { bubbles: true }));
                                        });
                                    } else {
                                        setTimeout(setupNameCapitalization, 100);
                                    }
                                }
                                setupNameCapitalization();
                            });
                            """
                        ),
                    ),
                    class_="col-md-6 mb-4",
                ),
                ui.tags.div(
                    ui.tags.label(
                        content["signature_label"], class_="form-label fw-medium mb-3"
                    ),
                    ui.tags.div(
                        ui.tags.canvas(
                            id="signature-canvas",
                            width="450",
                            height="180",
                            style="width: 100%; max-width: 450px; height: auto;",
                        ),
                        ui.tags.div(
                            ui.tags.button(
                                "✨ Clear Signature",
                                type="button",
                                id="clear-signature",
                                class_="btn btn-outline-secondary btn-sm me-3",
                            ),
                            ui.tags.span(
                                "✓ Signature captured",
                                id="signature-check",
                                style="color: #059669; font-weight: 600; display: none;",
                            ),
                            class_="mt-3 d-flex align-items-center",
                        ),
                        ui.tags.div(
                            "✍️ Sign here with your finger or mouse • Use landscape mode on mobile for best experience",
                            class_="text-muted small mt-2 text-center",
                        ),
                        class_="signature-section",
                    ),
                    ui.tags.div(
                        ui.input_text(
                            "signature_data",
                            None,
                            value="",
                        ),
                        style="display: none;",
                    ),
                    class_="col-md-6 mb-4",
                ),
                class_="row",
            ),
            ui.tags.div(
                ui.input_action_button(
                    "submit_waiver",
                    f"🚀 {content['submit_button']}",
                    class_="btn btn-primary btn-lg w-100",
                ),
                class_="d-grid gap-2",
            ),
            class_="form-section",
        ),
    )


app_ui = ui.page_fluid(
    ui.tags.style(
        """
        @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');
        
        body { 
            background: linear-gradient(135deg, #f5f7fa 0%, #c3cfe2 100%);
            min-height: 100vh;
            font-family: 'Inter', -apple-system, BlinkMacSystemFont, 'Segoe UI', 'PingFang SC', 'Hiragino Sans GB', 'Microsoft YaHei', system-ui, sans-serif;
            line-height: 1.6;
        }
        
        .waiver-container { 
            max-width: 900px; 
            margin: 2rem auto; 
            background: white; 
            border-radius: 20px; 
            box-shadow: 0 20px 60px rgba(0,0,0,0.15), 0 8px 25px rgba(0,0,0,0.08);
            overflow: hidden;
            border: 1px solid rgba(255,255,255,0.8);
            backdrop-filter: blur(10px);
        }
        
        .header-section {
            background: linear-gradient(135deg, #166534 0%, #15803d 100%);
            position: relative;
            overflow: hidden;
        }
        
        .form-content {
            padding: 3rem 2.5rem 2.5rem;
        }
        
        .form-control {
            border: 2px solid #e5e7eb;
            border-radius: 12px;
            padding: 0.875rem 1rem;
            font-size: 0.95rem;
            transition: all 0.3s ease;
            background: #fafafa;
        }
        
        .form-control:focus {
            border-color: #059669;
            box-shadow: 0 0 0 3px rgba(5, 150, 105, 0.1);
            background: white;
            transform: translateY(-1px);
        }
        
        .form-check-input {
            width: 1.25rem;
            height: 1.25rem;
            border: 2px solid #d1d5db;
            border-radius: 6px;
            background: #fafafa;
            transition: all 0.3s ease;
        }
        
        .form-check-input:focus {
            border-color: #059669;
            box-shadow: 0 0 0 3px rgba(5, 150, 105, 0.1);
        }
        
        .form-check-input:checked {
            background-color: #059669 !important;
            border-color: #059669 !important;
        }
        
        .form-label {
            font-weight: 600;
            color: #374151;
            margin-bottom: 0.5rem;
            font-size: 0.95rem;
        }
        
        .signature-section {
            background: #f8fafc;
            border: 2px dashed #cbd5e1;
            border-radius: 16px;
            padding: 1.5rem;
            transition: all 0.3s ease;
        }
        
        .signature-section:hover {
            border-color: #059669;
            background: #f0fdf4;
        }
        
        #signature-canvas {
            border: 2px solid #e2e8f0;
            border-radius: 12px;
            cursor: crosshair;
            touch-action: none;
            user-select: none;
            -webkit-user-select: none;
            -moz-user-select: none;
            -ms-user-select: none;
            width: 100%;
            max-width: 450px;
            min-height: 180px;
            background: white;
            box-shadow: 0 4px 12px rgba(0,0,0,0.08);
            transition: all 0.3s ease;
            position: relative;
            display: block;
        }
        
        #signature-canvas:hover {
            border-color: #059669;
            transform: translateY(-2px);
            box-shadow: 0 8px 24px rgba(0,0,0,0.12);
        }
        
        #signature-canvas:active,
        #signature-canvas:focus {
            border-color: #059669;
            outline: none;
        }
        
        @media (max-width: 768px) {
            #signature-canvas {
                min-height: 200px;
                max-width: 100%;
                border-width: 3px;
                border-radius: 16px;
                touch-action: manipulation;
            }
            
            .signature-section {
                padding: 1rem;
                border-width: 3px;
                touch-action: manipulation;
            }
            
            .signature-section .btn {
                padding: 0.875rem 1.5rem;
                font-size: 1rem;
                font-weight: 600;
                touch-action: manipulation;
            }
        }
        
        @media (pointer: coarse) {
            #signature-canvas {
                min-height: 220px;
                border-width: 3px;
            }
        }
        
        .btn {
            border-radius: 12px;
            padding: 0.75rem 1.5rem;
            font-weight: 600;
            font-size: 0.95rem;
            transition: all 0.3s ease;
            border: none;
        }
        
        .btn-primary {
            background: linear-gradient(135deg, #059669 0%, #10b981 100%);
            box-shadow: 0 4px 15px rgba(5, 150, 105, 0.4);
        }
        
        .btn-primary:hover {
            transform: translateY(-2px);
            box-shadow: 0 8px 25px rgba(5, 150, 105, 0.5);
            background: linear-gradient(135deg, #047857 0%, #059669 100%);
        }
        
        .btn-outline-secondary {
            border: 2px solid #e5e7eb;
            color: #6b7280;
            background: white;
        }
        
        .btn-outline-secondary:hover {
            background: #f9fafb;
            border-color: #d1d5db;
            transform: translateY(-1px);
        }
        
        .btn-outline-light {
            border: 2px solid rgba(255,255,255,0.3);
            color: white;
            background: rgba(255,255,255,0.1);
            backdrop-filter: blur(10px);
        }
        
        .btn-outline-light:hover {
            background: rgba(255,255,255,0.2);
            border-color: rgba(255,255,255,0.5);
            transform: translateY(-1px);
        }
        
        .success-message {
            background: linear-gradient(135deg, #ecfdf5 0%, #d1fae5 100%);
            border: 2px solid #a7f3d0;
            border-radius: 20px;
            padding: 3rem;
            text-align: center;
            margin: 2rem 0;
            box-shadow: 0 10px 30px rgba(5, 150, 105, 0.1);
        }
        
        .checkmark {
            width: 80px;
            height: 80px;
            border-radius: 50%;
            background: linear-gradient(135deg, #059669 0%, #10b981 100%);
            color: white;
            display: inline-flex;
            align-items: center;
            justify-content: center;
            font-size: 2.5rem;
            margin-bottom: 1.5rem;
            box-shadow: 0 8px 25px rgba(5, 150, 105, 0.3);
        }
        
        .waiver-title {
            font-size: 1.75rem;
            font-weight: 700;
            color: #1f2937;
            margin-bottom: 2rem;
            text-align: center;
            line-height: 1.3;
        }
        
        .waiver-content {
            font-size: 0.95rem;
            line-height: 1.7;
            color: #4b5563;
        }
        
        .waiver-content ol, .waiver-content ul {
            padding-left: 1.5rem;
        }
        
        .waiver-content li {
            margin-bottom: 0.75rem;
        }
        
        .logo-container {
            background: rgba(255,255,255,0.95);
            padding: 1rem;
            border-radius: 12px;
            box-shadow: 0 4px 15px rgba(0,0,0,0.1);
            margin-bottom: 1.5rem;
            backdrop-filter: blur(10px);
        }
        
        .divider {
            height: 2px;
            background: linear-gradient(90deg, transparent 0%, #e5e7eb 50%, transparent 100%);
            margin: 2.5rem 0;
            border: none;
        }
        
        .form-section {
            background: #fafafa;
            border-radius: 16px;
            padding: 2rem;
            margin: 2rem 0;
            border: 1px solid #e5e7eb;
        }
        
        .waiver-container * {
            -webkit-print-color-adjust: exact !important;
            color-adjust: exact !important;
        }
        
        @media (max-width: 768px) {
            .waiver-container {
                margin: 1rem;
                border-radius: 16px;
            }
            .form-content {
                padding: 2rem 1.5rem;
            }
            .waiver-title {
                font-size: 1.5rem;
            }
        }
        
        #submitting-overlay { 
            position: fixed; 
            inset: 0; 
            display: none; 
            align-items: center; 
            justify-content: center; 
            z-index: 2000; 
            background: rgba(0,0,0,0.6);
            backdrop-filter: blur(4px);
        }
        
        #submitting-overlay .inner { 
            color: #fff; 
            font-weight: 600; 
            font-size: 1.1rem; 
            display: flex; 
            align-items: center; 
            gap: 1rem;
            background: rgba(0,0,0,0.8);
            padding: 2rem 3rem;
            border-radius: 16px;
            box-shadow: 0 20px 60px rgba(0,0,0,0.3);
        }
    """
    ),
    ui.tags.script(
        """
        window.SignatureCanvas = window.SignatureCanvas || {
            canvas: null,
            ctx: null,
            isDrawing: false,
            lastX: 0,
            lastY: 0,
            isCanvasReady: false,
            
            init: function() {
                this.canvas = document.getElementById('signature-canvas');
                if (!this.canvas) {
                    setTimeout(() => this.init(), 100);
                    return;
                }
                
                this.setupCanvasContext();
                this.setupEventListeners();
                this.isCanvasReady = true;
            },
            
            reinit: function() {
                this.isCanvasReady = false;
                this.canvas = null;
                this.ctx = null;
                this.isDrawing = false;
                setTimeout(() => this.init(), 200);
            },
            
            setupCanvasContext: function() {
                this.ctx = this.canvas.getContext('2d');
                
                const resizeCanvas = () => {
                    const rect = this.canvas.getBoundingClientRect();
                    const dpr = window.devicePixelRatio || 1;
                    const isMobile = window.innerWidth <= 768;
                    
                    const currentData = this.isCanvasReady ? this.canvas.toDataURL() : null;
                    
                    this.canvas.width = rect.width * dpr;
                    this.canvas.height = (isMobile ? 200 : 180) * dpr;
                    
                    this.canvas.style.width = rect.width + 'px';
                    this.canvas.style.height = (isMobile ? 200 : 180) + 'px';
                    
                    this.ctx.scale(dpr, dpr);
                    this.ctx.strokeStyle = '#000';
                    this.ctx.lineWidth = isMobile ? 4 : 3;
                    this.ctx.lineCap = 'round';
                    this.ctx.lineJoin = 'round';
                    this.ctx.globalCompositeOperation = 'source-over';
                    
                    if (currentData && this.isCanvasReady) {
                        const img = new Image();
                        img.onload = () => {
                            this.ctx.drawImage(img, 0, 0, this.canvas.width / dpr, this.canvas.height / dpr);
                        };
                        img.src = currentData;
                    }
                };
                
                resizeCanvas();
                window.addEventListener('resize', this.debounce(resizeCanvas, 250));
            },
            
            setupEventListeners: function() {
                this.canvas.addEventListener('mousedown', (e) => this.startDrawing(e), { passive: false });
                this.canvas.addEventListener('mousemove', (e) => this.draw(e), { passive: false });
                this.canvas.addEventListener('mouseup', (e) => this.stopDrawing(e));
                this.canvas.addEventListener('mouseleave', (e) => this.stopDrawing(e));
                
                this.canvas.addEventListener('touchstart', (e) => this.handleTouchStart(e), { passive: false });
                this.canvas.addEventListener('touchmove', (e) => this.handleTouchMove(e), { passive: false });
                this.canvas.addEventListener('touchend', (e) => this.stopDrawing(e), { passive: false });
                this.canvas.addEventListener('touchcancel', (e) => this.stopDrawing(e), { passive: false });
                
                setTimeout(() => {
                    const clearBtn = document.getElementById('clear-signature');
                    if (clearBtn) {
                        clearBtn.onclick = () => this.clearCanvas();
                    }
                }, 100);
            },
            
            debounce: function(func, wait) {
                let timeout;
                return function executedFunction(...args) {
                    const later = () => {
                        clearTimeout(timeout);
                        func(...args);
                    };
                    clearTimeout(timeout);
                    timeout = setTimeout(later, wait);
                };
            },
            
            getPos: function(e) {
                if (!this.canvas) return { x: 0, y: 0 };
                const rect = this.canvas.getBoundingClientRect();
                const scaleX = this.canvas.width / rect.width;
                const scaleY = this.canvas.height / rect.height;
                
                const clientX = e.clientX || (e.touches && e.touches[0] && e.touches[0].clientX);
                const clientY = e.clientY || (e.touches && e.touches[0] && e.touches[0].clientY);
                
                return {
                    x: (clientX - rect.left) * scaleX / (window.devicePixelRatio || 1),
                    y: (clientY - rect.top) * scaleY / (window.devicePixelRatio || 1)
                };
            },
            
            startDrawing: function(e) {
                if (!this.ctx) return;
                e.preventDefault();
                this.isDrawing = true;
                const pos = this.getPos(e);
                this.lastX = pos.x;
                this.lastY = pos.y;
                
                this.ctx.beginPath();
                this.ctx.moveTo(this.lastX, this.lastY);
                this.ctx.lineTo(this.lastX, this.lastY);
                this.ctx.stroke();
            },
            
            draw: function(e) {
                if (!this.isDrawing || !this.ctx) return;
                e.preventDefault();
                
                const pos = this.getPos(e);
                this.ctx.beginPath();
                this.ctx.moveTo(this.lastX, this.lastY);
                this.ctx.lineTo(pos.x, pos.y);
                this.ctx.stroke();
                
                this.lastX = pos.x;
                this.lastY = pos.y;
                
                this.updateSignatureData();
            },
            
            stopDrawing: function(e) {
                if (this.isDrawing && this.ctx) {
                    e && e.preventDefault();
                    this.isDrawing = false;
                    this.ctx.beginPath();
                    this.updateSignatureData();
                }
            },
            
            handleTouchStart: function(e) {
                e.preventDefault();
                if (e.touches && e.touches.length === 1) {
                    const touch = e.touches[0];
                    this.startDrawing({ 
                        clientX: touch.clientX, 
                        clientY: touch.clientY,
                        preventDefault: () => {}
                    });
                }
            },
            
            handleTouchMove: function(e) {
                e.preventDefault();
                if (e.touches && e.touches.length === 1) {
                    const touch = e.touches[0];
                    this.draw({ 
                        clientX: touch.clientX, 
                        clientY: touch.clientY,
                        preventDefault: () => {}
                    });
                }
            },
            
            clearCanvas: function() {
                if (this.ctx && this.canvas) {
                    this.ctx.clearRect(0, 0, this.canvas.width, this.canvas.height);
                    this.updateSignatureData();
                }
            },
            
            updateSignatureData: function() {
                if (this.canvas && this.ctx) {
                    const dataURL = this.canvas.toDataURL('image/png');
                    const input = document.getElementById('signature_data');
                    if (input) {
                        input.value = dataURL;
                        const inputEvent = new Event('input', { bubbles: true });
                        const changeEvent = new Event('change', { bubbles: true });
                        input.dispatchEvent(inputEvent);
                        input.dispatchEvent(changeEvent);
                    }
                    
                    const isEmpty = this.isCanvasEmpty();
                    const checkmark = document.getElementById('signature-check');
                    if (checkmark) {
                        checkmark.style.display = isEmpty ? 'none' : 'inline';
                    }
                }
            },
            
            isCanvasEmpty: function() {
                if (!this.canvas || !this.ctx) return true;
                const blank = document.createElement('canvas');
                blank.width = this.canvas.width;
                blank.height = this.canvas.height;
                const blankCtx = blank.getContext('2d');
                blankCtx.fillStyle = 'white';
                blankCtx.fillRect(0, 0, blank.width, blank.height);
                return this.canvas.toDataURL() === blank.toDataURL();
            }
        };
        
        document.addEventListener('DOMContentLoaded', function() {
            window.SignatureCanvas.init();
            
            const observer = new MutationObserver(function(mutations) {
                mutations.forEach(function(mutation) {
                    if (mutation.type === 'childList') {
                        const canvas = document.getElementById('signature-canvas');
                        if (canvas && canvas !== window.SignatureCanvas.canvas) {
                            window.SignatureCanvas.reinit();
                        }
                    }
                });
            });
            
            observer.observe(document.body, {
                childList: true,
                subtree: true
            });
        });
        
        async function captureScreenshot(){
            console.log('[JS] captureScreenshot called');
            if(!window.html2canvas) {
                console.error('[JS] html2canvas not available');
                return;
            }
            try{
                const target = document.querySelector('.waiver-container') || document.body;
                console.log('[JS] Target element for capture:', target);
                window.scrollTo(0,0);
                
                const canvas = await html2canvas(target, {
                    scale: 1,
                    backgroundColor: '#ffffff',
                    useCORS:true,
                    logging:true,
                    allowTaint: false,
                    height: target.scrollHeight,
                    width: target.scrollWidth,
                    ignoreElements: el => {
                        const ignore = el.id==='page_screenshot' || el.id==='signature_data' || el.id==='submitting-overlay';
                        if (ignore) console.log('[JS] Ignoring element:', el.id);
                        return ignore;
                    }
                });
                
                console.log('[JS] Canvas created:', canvas.width, 'x', canvas.height);
                const dataURL = canvas.toDataURL('image/png');
                console.log('[JS] DataURL created, length:', dataURL.length);
                
                const input = document.getElementById('page_screenshot');
                console.log('[JS] Input element found:', !!input);
                if (input){
                    console.log('[JS] Current input value length:', input.value.length);
                    if (!input.value || input.value.length < 100) {
                        input.value = dataURL;
                        console.log('[JS] Set input value, new length:', input.value.length);
                        input.dispatchEvent(new Event('input',{bubbles:true}));
                        input.dispatchEvent(new Event('change',{bubbles:true}));
                        if(window.Shiny && window.Shiny.setInputValue){
                            console.log('[JS] Using Shiny.setInputValue');
                            Shiny.setInputValue('page_screenshot', dataURL, {priority:'event'});
                        }
                    } else {
                        console.log('[JS] Input already has screenshot data');
                    }
                } else {
                    console.error('[JS] page_screenshot input not found');
                }
            }catch(err){ 
                console.error('[JS] captureScreenshot error', err); 
            }
        }

        document.addEventListener('DOMContentLoaded', function() {
            setTimeout(function() {
                const submitButton = document.querySelector('button[id*="submit_waiver"]');
                if (submitButton) {
                    console.log('[JS] Found submit button, adding screenshot capture listener');
                    submitButton.addEventListener('click', function(event) {
                        console.log('[JS] Submit button clicked - capturing screenshot immediately');
                        const overlay = document.getElementById('submitting-overlay');
                        if (overlay) overlay.style.display = 'flex';
                        
                        setTimeout(() => {
                            console.log('[JS] Attempting immediate screenshot capture');
                            captureScreenshot().catch((err) => {
                                console.error('[JS] Immediate capture failed:', err);
                            });
                        }, 200);
                    }, true);
                } else {
                    console.log('[JS] Submit button not found');
                }
            }, 1000);
        });
        """
    ),
    ui.tags.script(
        """
        let screenshotMethods = {
            html2canvas: null,
            fallback: null
        };

        function loadHtml2Canvas() {
            return new Promise((resolve) => {
                if (window.html2canvas) {
                    screenshotMethods.html2canvas = window.html2canvas;
                    resolve(true);
                    return;
                }
                
                const script = document.createElement('script');
                script.src = 'https://html2canvas.hertzen.com/dist/html2canvas.min.js';
                script.onload = () => {
                    screenshotMethods.html2canvas = window.html2canvas;
                    console.log('[JS] html2canvas loaded successfully');
                    resolve(true);
                };
                script.onerror = () => {
                    console.log('[JS] html2canvas failed to load, will use fallback');
                    resolve(false);
                };
                document.head.appendChild(script);
            });
        }

        function captureWithCanvas() {
            return new Promise((resolve) => {
                try {
                    const waiver = document.querySelector('.waiver-container');
                    if (!waiver) {
                        resolve(null);
                        return;
                    }

                    const canvas = document.createElement('canvas');
                    const ctx = canvas.getContext('2d');
                    
                    canvas.width = waiver.scrollWidth;
                    canvas.height = waiver.scrollHeight;
                    
                    ctx.fillStyle = '#ffffff';
                    ctx.fillRect(0, 0, canvas.width, canvas.height);
                    
                    ctx.fillStyle = '#000000';
                    ctx.font = '14px Arial';
                    
                    const text = waiver.innerText || 'Waiver content captured';
                    const lines = text.split('\\n');
                    let y = 30;
                    
                    lines.forEach(line => {
                        if (y < canvas.height - 20) {
                            ctx.fillText(line.substring(0, 80), 20, y);
                            y += 20;
                        }
                    });
                    
                    const sigCanvas = document.getElementById('signature-canvas');
                    if (sigCanvas) {
                        try {
                            ctx.drawImage(sigCanvas, 20, y + 20);
                        } catch (e) {
                            console.log('[JS] Could not copy signature canvas');
                        }
                    }
                    
                    resolve(canvas.toDataURL('image/png'));
                } catch (error) {
                    console.error('[JS] Canvas fallback error:', error);
                    resolve(null);
                }
            });
        }

        function captureWithSVG() {
            return new Promise((resolve) => {
                try {
                    const waiver = document.querySelector('.waiver-container');
                    if (!waiver) {
                        resolve(null);
                        return;
                    }

                    const serializer = new XMLSerializer();
                    const rect = waiver.getBoundingClientRect();
                    
                    const svg = document.createElementNS('http://www.w3.org/2000/svg', 'svg');
                    svg.setAttribute('width', rect.width);
                    svg.setAttribute('height', rect.height);
                    
                    const foreignObject = document.createElementNS('http://www.w3.org/2000/svg', 'foreignObject');
                    foreignObject.setAttribute('width', '100%');
                    foreignObject.setAttribute('height', '100%');
                    
                    const clonedWaiver = waiver.cloneNode(true);
                    foreignObject.appendChild(clonedWaiver);
                    svg.appendChild(foreignObject);
                    
                    const svgString = serializer.serializeToString(svg);
                    const dataURL = 'data:image/svg+xml;charset=utf-8,' + encodeURIComponent(svgString);
                    
                    resolve(dataURL);
                } catch (error) {
                    console.error('[JS] SVG capture error:', error);
                    resolve(null);
                }
            });
        }

        async function captureScreenshot() {
            console.log('[JS] Starting enhanced screenshot capture...');
            
            await new Promise(resolve => setTimeout(resolve, 500));
            
            const nameInput = document.querySelector('input[id*="participant_name"]');
            const agreementCheckbox = document.querySelector('input[id*="agreement"]');
            
            if (nameInput && nameInput.value) {
                console.log('[JS] Name input value:', nameInput.value);
                nameInput.style.border = '1px solid #ced4da';
                nameInput.style.backgroundColor = 'white';
                nameInput.style.color = 'black';
            }
            
            if (agreementCheckbox) {
                console.log('[JS] Agreement checkbox checked:', agreementCheckbox.checked);
                if (agreementCheckbox.checked) {
                    agreementCheckbox.style.backgroundColor = '#0d6efd';
                    agreementCheckbox.style.borderColor = '#0d6efd';
                }
            }
            
            let result = null;
            
            // Try html2canvas first
            if (screenshotMethods.html2canvas) {
                try {
                    console.log('[JS] Attempting html2canvas capture...');
                    const target = document.querySelector('.waiver-container') || document.body;
                    window.scrollTo(0, 0);
                    
                    const canvas = await screenshotMethods.html2canvas(target, {
                        scale: 1.5,
                        backgroundColor: '#ffffff',
                        useCORS: true,
                        allowTaint: false,
                        logging: false,
                        height: target.scrollHeight,
                        width: target.scrollWidth,
                        onclone: function(clonedDoc) {
                            const clonedNameInput = clonedDoc.querySelector('input[id*="participant_name"]');
                            const clonedCheckbox = clonedDoc.querySelector('input[id*="agreement"]');
                            
                            if (clonedNameInput && nameInput && nameInput.value) {
                                clonedNameInput.value = nameInput.value;
                                clonedNameInput.setAttribute('value', nameInput.value);
                                clonedNameInput.style.color = 'black';
                                clonedNameInput.style.backgroundColor = 'white';
                                clonedNameInput.style.border = '1px solid #ced4da';
                            }
                            
                            if (clonedCheckbox && agreementCheckbox) {
                                clonedCheckbox.checked = agreementCheckbox.checked;
                                if (agreementCheckbox.checked) {
                                    clonedCheckbox.setAttribute('checked', 'checked');
                                    clonedCheckbox.style.backgroundColor = '#0d6efd';
                                    clonedCheckbox.style.borderColor = '#0d6efd';
                                }
                            }
                            
                            const allInputs = target.querySelectorAll('input, textarea, select');
                            const clonedInputs = clonedDoc.querySelectorAll('input, textarea, select');
                            
                            allInputs.forEach((input, index) => {
                                if (clonedInputs[index]) {
                                    clonedInputs[index].value = input.value;
                                    if (input.type === 'checkbox' || input.type === 'radio') {
                                        clonedInputs[index].checked = input.checked;
                                    }
                                }
                            });
                        },
                        ignoreElements: (el) => {
                            return el.id === 'page_screenshot' || 
                                   el.id === 'signature_data' || 
                                   el.id === 'submitting-overlay';
                        }
                    });
                    
                    result = canvas.toDataURL('image/png');
                    console.log('[JS] html2canvas capture successful, size:', result.length);
                } catch (error) {
                    console.error('[JS] html2canvas capture failed:', error);
                }
            }
            
            if (!result) {
                console.log('[JS] Attempting enhanced canvas fallback...');
                result = await captureWithEnhancedCanvas();
                if (result) {
                    console.log('[JS] Enhanced canvas fallback successful, size:', result.length);
                }
            }
            
            if (!result) {
                console.log('[JS] Attempting SVG fallback...');
                result = await captureWithSVG();
                if (result) {
                    console.log('[JS] SVG fallback successful, size:', result.length);
                }
            }
            
            if (result) {
                const input = document.getElementById('page_screenshot');
                if (input) {
                    input.value = result;
                    input.dispatchEvent(new Event('input', {bubbles: true}));
                    input.dispatchEvent(new Event('change', {bubbles: true}));
                    
                    if (window.Shiny && window.Shiny.setInputValue) {
                        Shiny.setInputValue('page_screenshot', result, {priority: 'event'});
                    }
                    
                    console.log('[JS] Screenshot saved to input field');
                } else {
                    console.error('[JS] page_screenshot input not found');
                }
            } else {
                console.error('[JS] All screenshot methods failed');
            }
            
            return result;
        }

        function captureWithEnhancedCanvas() {
            return new Promise((resolve) => {
                try {
                    const waiver = document.querySelector('.waiver-container');
                    if (!waiver) {
                        resolve(null);
                        return;
                    }

                    const canvas = document.createElement('canvas');
                    const ctx = canvas.getContext('2d');
                    
                    canvas.width = waiver.scrollWidth || 800;
                    canvas.height = waiver.scrollHeight || 1200;
                    
                    ctx.fillStyle = '#ffffff';
                    ctx.fillRect(0, 0, canvas.width, canvas.height);
                    
                    ctx.fillStyle = '#000000';
                    ctx.font = '16px Arial';
                    
                    let y = 50;
                    
                    ctx.font = 'bold 20px Arial';
                    ctx.fillText('BICYCLE PROGRAM WAIVER AND RELEASE FROM LIABILITY', 50, y);
                    y += 40;
                    
                    const nameInput = document.querySelector('input[id*="participant_name"]');
                    if (nameInput && nameInput.value) {
                        ctx.font = '16px Arial';
                        ctx.fillText('Participant Name: ' + nameInput.value, 50, y);
                        y += 30;
                    }
                    
                    const agreementCheckbox = document.querySelector('input[id*="agreement"]');
                    if (agreementCheckbox) {
                        ctx.fillText('Agreement: ' + (agreementCheckbox.checked ? '✓ Agreed' : '☐ Not Agreed'), 50, y);
                        y += 30;
                    }
                    
                    ctx.fillText('Date: ' + new Date().toLocaleDateString(), 50, y);
                    y += 40;
                    
                    const sigCanvas = document.getElementById('signature-canvas');
                    if (sigCanvas) {
                        try {
                            ctx.fillText('Signature:', 50, y);
                            y += 20;
                            ctx.drawImage(sigCanvas, 50, y);
                            y += 160;
                        } catch (e) {
                            ctx.fillText('Signature: [Signature Present]', 50, y);
                            y += 30;
                        }
                    }
                    
                    const textContent = waiver.innerText || '';
                    const lines = textContent.split('\n').filter(line => line.trim().length > 0);
                    ctx.font = '12px Arial';
                    
                    lines.slice(0, 50).forEach(line => {
                        if (y < canvas.height - 30) {
                            const words = line.split(' ');
                            let currentLine = '';
                            
                            words.forEach(word => {
                                const testLine = currentLine + word + ' ';
                                const metrics = ctx.measureText(testLine);
                                
                                if (metrics.width > canvas.width - 100) {
                                    ctx.fillText(currentLine, 50, y);
                                    currentLine = word + ' ';
                                    y += 15;
                                } else {
                                    currentLine = testLine;
                                }
                            });
                            
                            if (currentLine.trim().length > 0) {
                                ctx.fillText(currentLine, 50, y);
                                y += 15;
                            }
                        }
                    });
                    
                    resolve(canvas.toDataURL('image/png'));
                } catch (error) {
                    console.error('[JS] Enhanced canvas fallback error:', error);
                    resolve(null);
                }
            });
        }

        Shiny.addCustomMessageHandler('capture_page_screenshot', async () => {
            console.log('[JS] capture_page_screenshot handler called');
            await captureScreenshot();
        });
        
        Shiny.addCustomMessageHandler('hide_submitting_overlay', () => {
            const ov = document.getElementById('submitting-overlay');
            if (ov) ov.style.display = 'none';
        });
        
        Shiny.addCustomMessageHandler('clear_signature_canvas', () => {
            if (window.SignatureCanvas) {
                window.SignatureCanvas.clearCanvas();
            }
        });

        document.addEventListener('DOMContentLoaded', () => {
            loadHtml2Canvas();
        });
        """
    ),
    tags.div(
        ui.output_ui("header_section"),
        ui.tags.div(
            ui.output_ui("main_content"),
            class_="form-content",
        ),
        ui.tags.footer(
            "Hope's Corner, Inc. Bicycle Program Waiver",
            class_="text-center text-muted small mt-4 py-3",
        ),
        class_="waiver-container",
    ),
    ui.tags.div(
        ui.input_text(
            "page_screenshot",
            None,
            value="",
        ),
        style="display: none;",
    ),
    ui.tags.div(
        ui.tags.div(
            ui.tags.div(
                class_="spinner-border text-light spinner-border-sm", role="status"
            ),
            ui.tags.span("Submitting..."),
            class_="inner",
        ),
        id="submitting-overlay",
    ),
)


def server(input, output, session):
    language = reactive.Value("en")
    is_submitted = reactive.Value(False)
    submitting = reactive.Value(False)
    status_message = reactive.Value("")
    status_type = reactive.Value("info")

    AZURE_TENANT_ID = os.getenv("AZURE_TENANT_ID")
    AZURE_CLIENT_ID = os.getenv("AZURE_CLIENT_ID")
    AZURE_CLIENT_SECRET = os.getenv("AZURE_CLIENT_SECRET")
    SHAREPOINT_SITE_URL = os.getenv("SHAREPOINT_SITE_URL")
    SHAREPOINT_EXCEL_FILE_PATH = os.getenv("SHAREPOINT_EXCEL_FILE_PATH")
    SHAREPOINT_TABLE_NAME = os.getenv("SHAREPOINT_TABLE_NAME")

    def have_graph_config():
        return all(
            [
                AZURE_TENANT_ID,
                AZURE_CLIENT_ID,
                AZURE_CLIENT_SECRET,
                SHAREPOINT_SITE_URL,
                SHAREPOINT_EXCEL_FILE_PATH,
                SHAREPOINT_TABLE_NAME,
            ]
        )

    def graph_token():
        if not have_graph_config():
            return None
        try:
            url = (
                f"https://login.microsoftonline.com/{AZURE_TENANT_ID}/oauth2/v2.0/token"
            )
            data = {
                "client_id": AZURE_CLIENT_ID,
                "client_secret": AZURE_CLIENT_SECRET,
                "scope": "https://graph.microsoft.com/.default",
                "grant_type": "client_credentials",
            }
            resp = requests.post(url, data=data, timeout=15)
            resp.raise_for_status()
            return resp.json().get("access_token")
        except Exception as e:
            print(f"[graph] token error: {e}")
            return None

    def site_id(token: str):
        try:
            parsed = urlparse(SHAREPOINT_SITE_URL)
            hostname = parsed.netloc
            site_path = parsed.path.lstrip("/")
            url = f"https://graph.microsoft.com/v1.0/sites/{hostname}:/{site_path}?$select=id"
            r = requests.get(
                url, headers={"Authorization": f"Bearer {token}"}, timeout=15
            )
            r.raise_for_status()
            return r.json()["id"]
        except Exception as e:
            print(f"[graph] site id error: {e}")
            return None

    def ensure_excel_file_exists(token: str, sp_site_id: str):
        try:
            file_path_enc = SHAREPOINT_EXCEL_FILE_PATH.replace(" ", "%20")
            url = f"https://graph.microsoft.com/v1.0/sites/{sp_site_id}/drive/root:/{file_path_enc}"
            r = requests.get(
                url, headers={"Authorization": f"Bearer {token}"}, timeout=20
            )

            if r.status_code == 404:
                print("[graph] Excel file not found, creating new workbook")

                wb = Workbook()
                ws = wb.active
                ws.title = "WaiverData"

                headers = ["Name", "Date", "Language", "Timestamp", "Screenshot_File"]
                for idx, header in enumerate(headers, 1):
                    ws.cell(row=1, column=idx, value=header)

                sample_row = [
                    "Sample Name",
                    "2024-01-01",
                    "en",
                    "2024-01-01 12:00:00",
                    "Sample_Name_20240101_screenshot.png",
                ]
                for idx, value in enumerate(sample_row, 1):
                    ws.cell(row=2, column=idx, value=value)

                from openpyxl.worksheet.table import Table, TableStyleInfo

                tab = Table(displayName=SHAREPOINT_TABLE_NAME, ref="A1:E2")
                style = TableStyleInfo(
                    name="TableStyleMedium2",
                    showFirstColumn=False,
                    showLastColumn=False,
                    showRowStripes=True,
                    showColumnStripes=True,
                )
                tab.tableStyleInfo = style
                ws.add_table(tab)

                from io import BytesIO

                buffer = BytesIO()
                wb.save(buffer)
                buffer.seek(0)

                folder_path = "/".join(SHAREPOINT_EXCEL_FILE_PATH.split("/")[:-1])
                filename = SHAREPOINT_EXCEL_FILE_PATH.split("/")[-1]
                upload_url = f"https://graph.microsoft.com/v1.0/sites/{sp_site_id}/drive/root:/{folder_path.replace(' ', '%20')}/{filename}:/content"

                upload_resp = requests.put(
                    upload_url,
                    headers={"Authorization": f"Bearer {token}"},
                    data=buffer.getvalue(),
                    timeout=30,
                )
                if upload_resp.status_code < 300:
                    print("[graph] Excel file created successfully")
                    return True
                else:
                    print(
                        f"[graph] Failed to create Excel file: {upload_resp.status_code}"
                    )
                    return False
            else:
                print("[graph] Excel file already exists")
                return True
        except Exception as e:
            print(f"[graph] Error ensuring Excel file exists: {e}")
            return False

    def append_excel_row(token: str, sp_site_id: str, waiver_data: dict):
        try:
            if not ensure_excel_file_exists(token, sp_site_id):
                return False

            file_path_enc = SHAREPOINT_EXCEL_FILE_PATH.replace(" ", "%20")
            url = f"https://graph.microsoft.com/v1.0/sites/{sp_site_id}/drive/root:/{file_path_enc}:/workbook/tables/{SHAREPOINT_TABLE_NAME}/rows/add"

            date_obj = datetime.fromisoformat(waiver_data.get("date", ""))
            formatted_date = date_obj.strftime("%Y-%m-%d")

            row_values = [
                waiver_data.get("name", ""),
                formatted_date,
                waiver_data.get("language", ""),
                waiver_data.get("timestamp", ""),
                waiver_data.get("screenshot_filename", ""),
            ]

            body = {"values": [row_values]}
            r = requests.post(
                url,
                headers={
                    "Authorization": f"Bearer {token}",
                    "Content-Type": "application/json",
                },
                json=body,
                timeout=20,
            )

            if r.status_code < 300:
                print("[graph] Excel row appended successfully")
                return True
            else:
                session_url = f"https://graph.microsoft.com/v1.0/sites/{sp_site_id}/drive/root:/{file_path_enc}:/workbook/createSession"
                session_resp = requests.post(
                    session_url,
                    headers={
                        "Authorization": f"Bearer {token}",
                        "Content-Type": "application/json",
                    },
                    json={"persistChanges": True},
                    timeout=20,
                )

                if session_resp.status_code < 300:
                    session_id = session_resp.json().get("id")
                    print(f"[graph] Created workbook session: {session_id[:8]}...")

                    table_url = f"https://graph.microsoft.com/v1.0/sites/{sp_site_id}/drive/root:/{file_path_enc}:/workbook/tables/{SHAREPOINT_TABLE_NAME}/rows/add"
                    table_resp = requests.post(
                        table_url,
                        headers={
                            "Authorization": f"Bearer {token}",
                            "Content-Type": "application/json",
                            "workbook-session-id": session_id,
                        },
                        json=body,
                        timeout=20,
                    )

                    if table_resp.status_code < 300:
                        print("[graph] Excel row appended successfully using session")
                        return True
                    else:
                        print(
                            f"[graph] Failed to append Excel row with session: {table_resp.status_code} - {table_resp.text[:200]}"
                        )
                        return False
                else:
                    print(
                        f"[graph] Failed to create workbook session: {session_resp.status_code} - {session_resp.text[:200]}"
                    )
                    return False

                print(
                    f"[graph] Failed to append Excel row: {r.status_code} - {r.text[:200]}"
                )
                return False
        except Exception as e:
            print(f"[graph] Excel append error: {e}")
            return False

    def create_waiver_screenshot(waiver_data, content):
        """Create a high-quality server-side screenshot/image of the waiver data"""
        try:
            width = 1200
            height = 1800

            img = Image.new("RGB", (width, height), "#ffffff")
            draw = ImageDraw.Draw(img)

            def get_font_for_language(language, size):
                script_dir = os.path.dirname(os.path.abspath(__file__))
                fonts_dir = os.path.join(script_dir, "fonts")
                
                if language == "zh":
                    bundled_chinese_fonts = [
                        os.path.join(fonts_dir, "NotoSansCJKsc-Regular.otf"),
                        os.path.join(fonts_dir, "NotoSansCJKsc-Bold.otf")
                    ]
                    for font_path in bundled_chinese_fonts:
                        try:
                            if os.path.exists(font_path):
                                return ImageFont.truetype(font_path, size)
                        except:
                            continue
                
                bundled_western_fonts = [
                    os.path.join(fonts_dir, "OpenSans-Regular.ttf")
                ]
                for font_path in bundled_western_fonts:
                    try:
                        if os.path.exists(font_path):
                            return ImageFont.truetype(font_path, size)
                    except:
                        continue
                
                system_fonts = [
                    "/System/Library/Fonts/Helvetica.ttc",
                    "/System/Library/Fonts/Arial.ttf", 
                    "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
                    "/Windows/Fonts/arial.ttf"
                ]
                for font_path in system_fonts:
                    try:
                        return ImageFont.truetype(font_path, size)
                    except:
                        continue
                
                try:
                    return ImageFont.load_default()
                except:
                    return None

            current_language = waiver_data.get('language', 'en')
            title_font = get_font_for_language(current_language, 28)
            header_font = get_font_for_language(current_language, 22)
            normal_font = get_font_for_language(current_language, 16)
            small_font = get_font_for_language(current_language, 14)

            margin = 60
            y = 50

            header_height = 120
            draw.rectangle([(0, 0), (width, header_height)], fill="#059669")
            
            logo_text = "HOPE'S CORNER"
            subtitle_text = "SHARING MEALS, BUILDING COMMUNITY"
            if title_font:
                logo_bbox = draw.textbbox((0, 0), logo_text, font=title_font)
                logo_width = logo_bbox[2] - logo_bbox[0]
                logo_x = (width - logo_width) // 2
                draw.text((logo_x, 25), logo_text, fill="white", font=title_font)
                
                if normal_font:
                    subtitle_bbox = draw.textbbox((0, 0), subtitle_text, font=normal_font)
                    subtitle_width = subtitle_bbox[2] - subtitle_bbox[0]
                    subtitle_x = (width - subtitle_width) // 2
                    draw.text((subtitle_x, 65), subtitle_text, fill="white", font=normal_font)

            y = header_height + 40

            def safe_draw_text(draw, pos, text, font, fill="#000000", max_width=None):
                try:
                    if font and text:
                        if max_width and current_language == "zh":
                            lines = []
                            current_line = ""
                            for char in text:
                                test_line = current_line + char
                                try:
                                    bbox = draw.textbbox((0, 0), test_line, font=font)
                                    line_width = bbox[2] - bbox[0]
                                    if line_width > max_width and current_line:
                                        lines.append(current_line)
                                        current_line = char
                                    else:
                                        current_line = test_line
                                except:
                                    current_line = test_line
                            if current_line:
                                lines.append(current_line)
                            
                            x, y = pos
                            for line in lines[:6]:
                                try:
                                    draw.text((x, y), line, fill=fill, font=font)
                                    y += font.size + 5 if hasattr(font, 'size') else 25
                                except:
                                    draw.text((x, y), line.encode('utf-8', 'ignore').decode('utf-8'), fill=fill, font=font)
                                    y += font.size + 5 if hasattr(font, 'size') else 25
                            return y
                        else:
                            draw.text(pos, text, fill=fill, font=font)
                            return pos[1] + (font.size + 5 if hasattr(font, 'size') else 25)
                    else:
                        draw.text(pos, text[:100], fill=fill)
                        return pos[1] + 25
                except Exception as e:
                    try:
                        safe_text = text.encode('utf-8', 'ignore').decode('utf-8')
                        draw.text(pos, safe_text[:100], fill=fill)
                    except:
                        draw.text(pos, "[Text rendering error]", fill=fill)
                    return pos[1] + 25

            title_text = content["title"]
            if current_language == "zh":
                y = safe_draw_text(draw, (margin, y), title_text, title_font, "#1f2937", width - 2*margin)
            else:
                title_lines = textwrap.wrap(title_text, width=55)
                for line in title_lines:
                    if title_font:
                        try:
                            line_bbox = draw.textbbox((0, 0), line, font=title_font)
                            line_width = line_bbox[2] - line_bbox[0]
                            line_x = (width - line_width) // 2
                            draw.text((line_x, y), line, fill="#1f2937", font=title_font)
                        except:
                            draw.text((margin, y), line, fill="#1f2937")
                    else:
                        draw.text((margin, y), line, fill="#1f2937")
                    y += 35

            y += 30

            y = safe_draw_text(draw, (margin, y), content["logo_text1"], header_font or title_font, "black")
            y = safe_draw_text(draw, (margin, y), content["logo_subtext"], normal_font or title_font, "black")
            y += 10

            details_text = "FORM SUBMISSION DETAILS:" if current_language == "en" else "表单提交详情:" if current_language == "zh" else "DETALLES DE ENVÍO:"
            y = safe_draw_text(draw, (margin, y), details_text, header_font or title_font, "black")
            y += 5

            name_label = "Participant Name:" if current_language == "en" else "参与者姓名:" if current_language == "zh" else "Nombre del Participante:"
            name_text = f"{name_label} {waiver_data.get('name', 'N/A')}"
            y = safe_draw_text(draw, (margin, y), name_text, normal_font or title_font, "black")

            agreement_label = "Agreement Status:" if current_language == "en" else "协议状态:" if current_language == "zh" else "Estado del Acuerdo:"
            agreement_status = "✓ AGREED" if current_language == "en" else "✓ 已同意" if current_language == "zh" else "✓ ACEPTADO"
            not_agreed = "☐ NOT AGREED" if current_language == "en" else "☐ 未同意" if current_language == "zh" else "☐ NO ACEPTADO"
            agreement_text = f"{agreement_label} {agreement_status if waiver_data.get('agreement', False) else not_agreed}"
            y = safe_draw_text(draw, (margin, y), agreement_text, normal_font or title_font, "black")

            date_label = "Date:" if current_language == "en" else "日期:" if current_language == "zh" else "Fecha:"
            date_text = f"{date_label} {waiver_data.get('timestamp', 'N/A')}"
            y = safe_draw_text(draw, (margin, y), date_text, normal_font or title_font, "black")

            lang_mapping = {'en': 'English', 'es': 'Spanish', 'zh': 'Chinese'}
            lang_label = "Language:" if current_language == "en" else "语言:" if current_language == "zh" else "Idioma:"
            lang_text = f"{lang_label} {lang_mapping.get(waiver_data.get('language'), 'Unknown')}"
            y = safe_draw_text(draw, (margin, y), lang_text, normal_font or title_font, "black")
            y += 10

            signature_label = "SIGNATURE:" if current_language == "en" else "签名:" if current_language == "zh" else "FIRMA:"
            y = safe_draw_text(draw, (margin, y), signature_label, header_font or title_font, "black")

            signature_data = waiver_data.get("signature", "")
            if signature_data and signature_data.startswith("data:image"):
                try:
                    header, encoded = signature_data.split(",", 1)
                    signature_bytes = base64.b64decode(encoded)
                    signature_img = Image.open(io.BytesIO(signature_bytes))

                    if signature_img.mode != "RGBA":
                        signature_img = signature_img.convert("RGBA")

                    sig_width = min(400, signature_img.width)
                    sig_height = int(
                        signature_img.height * (sig_width / signature_img.width)
                    )
                    signature_img = signature_img.resize(
                        (sig_width, sig_height), Image.Resampling.LANCZOS
                    )

                    sig_background = Image.new("RGB", (sig_width, sig_height), "white")

                    if signature_img.mode == "RGBA":
                        sig_background.paste(signature_img, (0, 0), signature_img)
                    else:
                        sig_background.paste(signature_img, (0, 0))

                    img.paste(sig_background, (margin, y))
                    y += sig_height + 10
                    print(
                        f"[screenshot] Successfully processed signature: {sig_width}x{sig_height}"
                    )
                except Exception as e:
                    print(f"[screenshot] Could not process signature: {e}")
                    import traceback
                    traceback.print_exc()
                    no_display_text = "[Signature Present - Could not display]" if current_language == "en" else "[签名存在 - 无法显示]" if current_language == "zh" else "[Firma presente - No se puede mostrar]"
                    y = safe_draw_text(draw, (margin, y), no_display_text, normal_font or title_font, "black")
            else:
                no_signature_text = "[No signature provided]" if current_language == "en" else "[未提供签名]" if current_language == "zh" else "[No se proporcionó firma]"
                y = safe_draw_text(draw, (margin, y), no_signature_text, normal_font or title_font, "black")

            y += 20

            waiver_content_label = "WAIVER CONTENT:" if current_language == "en" else "弃权书内容:" if current_language == "zh" else "CONTENIDO DE RENUNCIA:"
            y = safe_draw_text(draw, (margin, y), waiver_content_label, header_font or title_font, "black")
            y += 5

            if current_language == "zh":
                y = safe_draw_text(draw, (margin, y), content["intro"], small_font or normal_font, "black", width - 2*margin - 20)
                y += 10
                
                for i, point in enumerate(content["points"][:3], 1):
                    point_text = f"{i}. {point}"
                    y = safe_draw_text(draw, (margin, y), point_text, small_font or normal_font, "black", width - 2*margin - 20)
                    if y > height - 200:
                        break
                
                if y < height - 150:
                    sub_points = content.get("sub_points", [])
                    for i, sub_point in enumerate(sub_points[:2], 1):
                        sub_text = f"  • {sub_point}"
                        y = safe_draw_text(draw, (margin + 20, y), sub_text, small_font or normal_font, "black", width - 2*margin - 40)
                        if y > height - 100:
                            break
            else:
                intro_lines = textwrap.wrap(content["intro"], width=90)
                for line in intro_lines[:12]:
                    y = safe_draw_text(draw, (margin, y), line, small_font or normal_font, "black")
                    if y > height - 150:
                        break

                if y < height - 300:
                    for i, point in enumerate(content["points"][:4], 1):
                        point_lines = textwrap.wrap(f"{i}. {point}", width=85)
                        for line in point_lines[:4]:
                            y = safe_draw_text(draw, (margin, y), line, small_font or normal_font, "black")
                            if y > height - 200:
                                break
                        if y > height - 200:
                            break

                if y < height - 250:
                    sub_points = content.get("sub_points", [])
                    for i, sub_point in enumerate(sub_points[:3], 1):
                        sub_lines = textwrap.wrap(f"  • {sub_point}", width=80)
                        for line in sub_lines[:3]:
                            y = safe_draw_text(draw, (margin + 20, y), line, small_font or normal_font, "black")
                            if y > height - 150:
                                break
                        if y > height - 150:
                            break

            if y < height - 120:
                y += 15
                important_label = "IMPORTANT:" if current_language == "en" else "重要:" if current_language == "zh" else "IMPORTANTE:"
                y = safe_draw_text(draw, (margin, y), important_label, header_font or title_font, "black")
                
                if current_language == "zh":
                    y = safe_draw_text(draw, (margin, y), content["final_agreement"], small_font or normal_font, "black", width - 2*margin - 20)
                else:
                    final_lines = textwrap.wrap(content["final_agreement"], width=85)
                    for line in final_lines[:8]:
                        y = safe_draw_text(draw, (margin, y), line, small_font or normal_font, "black")
                        if y > height - 80:
                            break

            buffer = io.BytesIO()
            img.save(buffer, format="PNG")
            buffer.seek(0)
            img_base64 = base64.b64encode(buffer.getvalue()).decode()

            return f"data:image/png;base64,{img_base64}"

        except Exception as e:
            print(f"[screenshot] Error creating waiver screenshot: {e}")
            return None

    def upload_screenshot_with_participant_name(
        token: str, sp_site_id: str, b64data: str, participant_name: str, timestamp: str
    ):
        print("[graph] Starting screenshot upload function")
        print(f"[graph] Data validation - b64data exists: {bool(b64data)}")
        print(
            f"[graph] Data validation - starts with data:image: {b64data.startswith('data:image') if b64data else False}"
        )
        print(f"[graph] Data validation - participant name: '{participant_name}'")
        print(
            f"[graph] Data validation - site_id: '{sp_site_id[:20] if sp_site_id else 'None'}...'"
        )

        if not (b64data and b64data.startswith("data:image")):
            print(
                f"[graph] Screenshot upload failed validation - b64data: {bool(b64data)}, starts with data:image: {b64data.startswith('data:image') if b64data else False}"
            )
            return False
        try:
            import re

            clean_name = re.sub(r"[^\w\s-]", "", participant_name.strip())
            clean_name = re.sub(r"[-\s]+", "_", clean_name)

            folder_path = (
                "/".join(SHAREPOINT_EXCEL_FILE_PATH.split("/")[:-1])
                or "Shared Documents"
            )

            print(f"[graph] SHAREPOINT_EXCEL_FILE_PATH: '{SHAREPOINT_EXCEL_FILE_PATH}'")
            print(f"[graph] Calculated folder_path: '{folder_path}'")

            screenshots_folder = f"{folder_path}/screenshots"
            print(f"[graph] Screenshots folder path: '{screenshots_folder}'")

            create_folder_url = f"https://graph.microsoft.com/v1.0/sites/{sp_site_id}/drive/root:/{folder_path.replace(' ', '%20')}:/children"
            folder_data = {
                "name": "screenshots",
                "folder": {},
                "@microsoft.graph.conflictBehavior": "replace",
            }

            folder_resp = requests.post(
                create_folder_url,
                headers={
                    "Authorization": f"Bearer {token}",
                    "Content-Type": "application/json",
                },
                json=folder_data,
                timeout=20,
            )

            if folder_resp.status_code < 300 or folder_resp.status_code == 409:
                print("[graph] Screenshots folder ready for use")
            else:
                print(
                    f"[graph] Screenshots folder creation response: {folder_resp.status_code} - {folder_resp.text[:100]}"
                )

            date_str = datetime.now().strftime("%Y%m%d")
            filename = f"{clean_name}_{date_str}_screenshot.png"

            file_api = f"https://graph.microsoft.com/v1.0/sites/{sp_site_id}/drive/root:/{screenshots_folder.replace(' ', '%20')}/{filename}:/content"

            print(f"[graph] Upload URL: '{file_api}'")

            try:
                header, encoded = b64data.split(",", 1)
                binary = base64.b64decode(encoded)
                print(
                    f"[graph] Successfully decoded base64 data, binary size: {len(binary)} bytes"
                )
            except Exception as decode_error:
                print(f"[graph] Failed to decode base64 data: {decode_error}")
                return False

            print(
                f"[graph] Uploading screenshot to SharePoint: {filename} (size: {len(binary)} bytes)"
            )
            print(f"[graph] Token length: {len(token)} characters")

            try:
                r = requests.put(
                    file_api,
                    headers={"Authorization": f"Bearer {token}"},
                    data=binary,
                    timeout=30,
                )
                print(f"[graph] Upload request completed with status: {r.status_code}")
            except Exception as upload_error:
                print(f"[graph] Upload request failed with exception: {upload_error}")
                return False

            if r.status_code < 300:
                print(f"[graph] Screenshot uploaded: {filename}")
                return f"{screenshots_folder}/{filename}"
            else:
                print(
                    f"[graph] Screenshot upload failed {r.status_code}: {r.text[:120]}"
                )
                return False
        except Exception as e:
            print(f"[graph] Screenshot upload error: {e}")
            return False

    @render.ui
    def header_section():
        content = waiver_content[language.get()]
        return create_header(content, language.get())

    @render.ui
    def main_content():
        if is_submitted.get():
            content = waiver_content[language.get()]
            submit_another_text = "Submit Another Waiver" if language.get() == "en" else "Enviar Otra Renuncia" if language.get() == "es" else "提交另一份弃权书"
            
            return ui.tags.div(
                ui.tags.div(
                    ui.tags.div("✓", class_="checkmark"),
                    ui.tags.h3(
                        content["success_message"], class_="h4 fw-bold text-success"
                    ),
                    ui.tags.div(
                        ui.input_action_button(
                            "submit_another",
                            submit_another_text,
                            class_="btn btn-primary btn-lg mt-3",
                        ),
                        class_="d-grid gap-2 mt-4",
                    ),
                    class_="success-message",
                ),
                ui.tags.script(
                    """
                    setTimeout(function() {
                        const overlay = document.getElementById('submitting-overlay');
                        if (overlay) overlay.style.display = 'none';
                    }, 100);
                    """
                ),
                class_="p-4",
            )
        else:
            content = waiver_content[language.get()]
            alert = None
            if status_message.get():
                alert = ui.tags.div(
                    status_message.get(),
                    class_=f"alert alert-{status_type.get()} py-2 px-3 mb-3",
                    role="alert",
                )
            return tags.div(
                alert if alert else "",
                create_waiver_content(content),
                create_form_section(content),
                class_="p-4",
            )

    @render.text
    def current_date_display():
        content = waiver_content[language.get()]
        today = datetime.now()
        if language.get() == "en":
            formatted_date = today.strftime("%B %d, %Y")
        elif language.get() == "es":
            months = [
                "enero",
                "febrero",
                "marzo",
                "abril",
                "mayo",
                "junio",
                "julio",
                "agosto",
                "septiembre",
                "octubre",
                "noviembre",
                "diciembre",
            ]
            formatted_date = f"{today.day} de {months[today.month-1]} de {today.year}"
        else:  # Chinese
            formatted_date = f"{today.year}年{today.month}月{today.day}日"

        return f"{content['date_label']}: {formatted_date}"

    @reactive.Effect
    @reactive.event(input.submit_another)
    def reset_for_new_waiver():
        is_submitted.set(False)
        status_message.set("")
        status_type.set("info")
        ui.update_text("participant_name", value="")
        ui.update_checkbox("agreement", value=False)
        ui.update_text("signature_data", value="")
        ui.update_action_button("submit_waiver", disabled=True)
        
        session.send_custom_message("clear_signature_canvas", {})

    @reactive.Effect
    @reactive.event(input.language_switch)
    def toggle_language():
        current_lang = language.get()
        if current_lang == "en":
            new_lang = "es"
        elif current_lang == "es":
            new_lang = "zh"
        else:
            new_lang = "en"
        language.set(new_lang)

    @reactive.Effect
    def update_submit_button():
        if (
            hasattr(input, "agreement")
            and hasattr(input, "participant_name")
            and hasattr(input, "signature_data")
        ):
            if (
                input.agreement()
                and input.participant_name()
                and input.signature_data()
            ):
                ui.update_action_button("submit_waiver", disabled=False)
            else:
                ui.update_action_button("submit_waiver", disabled=True)

    @reactive.Effect
    async def inject_html2canvas():
        await session.send_custom_message("inject_html2canvas", {})

    @reactive.Effect
    @reactive.event(input.submit_waiver)
    async def submit_waiver():
        if not (
            input.agreement() and input.participant_name() and input.signature_data()
        ):
            await session.send_custom_message("hide_submitting_overlay", {})
            return
        if submitting.get():
            await session.send_custom_message("hide_submitting_overlay", {})
            return

        print("[submit] Starting waiver submission...")

        submitting.set(True)
        ui.update_action_button(
            "submit_waiver",
            label=("Submitting..." if language.get() == "en" else "Enviando..."),
            disabled=True,
        )

        today = datetime.now()
        screenshot_data = ""

        print("[submit] Generating server-side screenshot...")

        current_lang = language.get()
        current_content = waiver_content[current_lang]

        participant_name = input.participant_name().strip()
        capitalized_name = ' '.join(word.capitalize() for word in participant_name.split())
        
        waiver_form_data = {
            "name": capitalized_name,
            "agreement": True,
            "signature": input.signature_data(),
            "timestamp": today.strftime("%Y-%m-%d %H:%M:%S"),
            "language": language.get(),
        }

        screenshot_data = create_waiver_screenshot(waiver_form_data, current_content)

        if screenshot_data:
            print(
                f"[submit] Generated server-side screenshot: {len(screenshot_data)} bytes"
            )
        else:
            print("[submit] Failed to generate server-side screenshot")
            screenshot_data = None

        import re

        clean_name = re.sub(r"[^\w\s-]", "", capitalized_name)
        clean_name = re.sub(r"[-\s]+", "_", clean_name)
        date_str = today.strftime("%Y%m%d")
        screenshot_filename = f"{clean_name}_{date_str}_screenshot.png"

        waiver_data = {
            "name": capitalized_name,
            "signature": input.signature_data(),
            "date": today.isoformat(),
            "language": language.get(),
            "timestamp": today.strftime("%Y-%m-%d %H:%M:%S"),
            "screenshot": screenshot_data,
            "screenshot_filename": screenshot_filename,
        }

        if not have_graph_config():
            print("[graph] SharePoint configuration is missing")
            status_message.set(
                "SharePoint configuration error. Cannot save waiver."
                if language.get() == "en"
                else "Error de configuración de SharePoint. No se puede guardar la renuncia."
            )
            status_type.set("danger")
            is_submitted.set(False)
            submitting.set(False)
            await session.send_custom_message("hide_submitting_overlay", {})
            return

        if have_graph_config():
            try:
                token = graph_token()
                if token:
                    sid = site_id(token)
                    if sid:
                        excel_success = append_excel_row(token, sid, waiver_data)
                        if not excel_success:
                            print("[graph] Failed to append data to Excel.")
                            status_message.set(
                                "Failed to save waiver to SharePoint Excel."
                                if language.get() == "en"
                                else "Error al guardar la renuncia en Excel de SharePoint."
                            )
                            status_type.set("danger")
                        else:
                            print("[graph] Successfully appended data to Excel.")

                        if waiver_data.get("screenshot"):
                            print(
                                f"[graph] About to upload screenshot for {waiver_data['name']}, data size: {len(waiver_data['screenshot'])}"
                            )
                            print(
                                f"[graph] Screenshot data preview: {waiver_data['screenshot'][:50]}..."
                            )
                            screenshot_path = upload_screenshot_with_participant_name(
                                token,
                                sid,
                                waiver_data["screenshot"],
                                waiver_data["name"],
                                waiver_data["timestamp"],
                            )
                            if screenshot_path:
                                print(f"[graph] Screenshot saved to: {screenshot_path}")
                            else:
                                print(
                                    "[graph] Failed to upload screenshot to SharePoint"
                                )
                        else:
                            print("[graph] No screenshot data to upload")
                            print(
                                f"[graph] Waiver data keys: {list(waiver_data.keys())}"
                            )
                            if waiver_data.get("screenshot") == "":
                                print("[graph] Screenshot data is empty string")
                    else:
                        print("[graph] Failed to get SharePoint site ID")
                        status_message.set(
                            "Failed to connect to SharePoint site."
                            if language.get() == "en"
                            else "Error al conectar con el sitio de SharePoint."
                        )
                        status_type.set("danger")
                        await session.send_custom_message("hide_submitting_overlay", {})
                else:
                    print("[graph] Failed to get Microsoft Graph token")
                    status_message.set(
                        "SharePoint authentication failed."
                        if language.get() == "en"
                        else "Error de autenticación con SharePoint."
                    )
                    status_type.set("danger")
                    await session.send_custom_message("hide_submitting_overlay", {})
            except Exception as e:
                print(f"[graph] SharePoint submission error: {e}")
                status_message.set(
                    "Error connecting to SharePoint."
                    if language.get() == "en"
                    else "Error al conectar con SharePoint."
                )
                status_type.set("danger")
                await session.send_custom_message("hide_submitting_overlay", {})

        is_submitted.set(True)
        submitting.set(False)
        status_message.set(
            "Waiver submitted successfully!"
            if language.get() == "en"
            else "¡Renuncia enviada con éxito!"
        )
        status_type.set("success")
        print(f"Waiver submitted - Participant: {waiver_data['name']}")
        await session.send_custom_message("hide_submitting_overlay", {})


app = App(app_ui, server)
